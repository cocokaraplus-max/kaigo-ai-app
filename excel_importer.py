"""
TASUKARU 既存Excel取り込みモジュール v2

機能:
  - アップロードされた既存のモニタリングxlsxを解析
  - シート名から最新月のモニタリング報告書を自動判定
  - 最新月で空欄の項目は、過去月のシートから値を拾う（フォールバック）
  - ケース記録シートから利用曜日を抽出
  - 目標・支援内容・基本情報を抽出して、patient_care_plansにインポート
"""
import re
from io import BytesIO
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook


# ========================================================
# シート名から年月を抽出
# ========================================================

def parse_reiwa_sheet_name(sheet_name: str) -> tuple:
    """
    "モニタリング報告書R8.4" → (8, 4)
    "ケース記録R7.12" → (7, 12)
    """
    m = re.search(r'R(\d+)\.(\d+)', sheet_name)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    m = re.search(r'(\d{4})[\.\-/](\d{1,2})', sheet_name)
    if m:
        y = int(m.group(1))
        era_year = y - 2018
        if 1 <= era_year <= 99:
            return (era_year, int(m.group(2)))
    return None


def list_monitoring_sheets_newest_first(workbook) -> list:
    """モニタリング報告書シートを新しい順に並べて返す [(era_year, month, sheet_name), ...]"""
    sheets = []
    for name in workbook.sheetnames:
        if "モニタリング" not in name:
            continue
        ym = parse_reiwa_sheet_name(name)
        if ym:
            sheets.append((ym[0], ym[1], name))
    sheets.sort(reverse=True)
    return sheets


def list_case_sheets_newest_first(workbook) -> list:
    """ケース記録シートを新しい順に並べて返す"""
    sheets = []
    for name in workbook.sheetnames:
        if "ケース記録" not in name:
            continue
        ym = parse_reiwa_sheet_name(name)
        if ym:
            sheets.append((ym[0], ym[1], name))
    sheets.sort(reverse=True)
    return sheets


# ========================================================
# セル値の抽出
# ========================================================

EXTRACTION_MAP = {
    # 基本情報
    "user_name_kana":       {"cell": "D11", "type": "string"},
    "user_name":            {"cell": "D12", "type": "string"},
    "gender":               {"cell": "H12", "type": "string"},
    "birthday":             {"cell": "I12", "type": "date"},
    "age":                  {"cell": "M12", "type": "number"},
    "care_level_text":      {"cell": "O12", "type": "string"},
    "care_level_num":       {"cell": "Q12", "type": "string"},
    "chart_staff_name":     {"cell": "R12", "type": "string"},
    # 短期目標
    "short_goal_period_from":   {"cell": "O14", "type": "date"},
    "short_goal_period_to":     {"cell": "S14", "type": "date"},
    "short_goal_function":      {"cell": "D15", "type": "string"},
    "short_goal_activity":      {"cell": "D16", "type": "string"},
    "short_goal_participation": {"cell": "D17", "type": "string"},
    # 長期目標
    "long_goal_period_from":   {"cell": "O18", "type": "date"},
    "long_goal_period_to":     {"cell": "S18", "type": "date"},
    "long_goal_function":      {"cell": "D19", "type": "string"},
    "long_goal_activity":      {"cell": "D20", "type": "string"},
    "long_goal_participation": {"cell": "D21", "type": "string"},
    # 支援内容
    "support_content_1": {"cell": "C32", "type": "string"},
    "support_content_2": {"cell": "C33", "type": "string"},
    "support_content_3": {"cell": "C34", "type": "string"},
    "support_content_4": {"cell": "C35", "type": "string"},
}

# ケース記録からの追加抽出
CASE_EXTRACTION_MAP = {
    "usage_weekday": {"cell": "B51", "type": "string"},  # "火・午前" など
}


def _convert_value(val, type_hint: str):
    if val is None:
        return None
    
    if type_hint == "date":
        if isinstance(val, datetime):
            return val.date().isoformat()
        if isinstance(val, date):
            return val.isoformat()
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                continue
        return None
    
    if type_hint == "number":
        try:
            if isinstance(val, (int, float)):
                return int(val)
            return int(str(val).strip())
        except (ValueError, TypeError):
            return None
    
    s = str(val).strip()
    return s if s else None


def extract_from_worksheet(ws, extraction_map) -> dict:
    result = {}
    for key, spec in extraction_map.items():
        val = ws[spec["cell"]].value
        result[key] = _convert_value(val, spec["type"])
    return result


def extract_with_fallback(workbook, monitoring_sheets: list) -> tuple:
    """
    最新月から順にシートを見ていき、各項目の値を拾う。
    最新月で空欄の項目は、次に新しい月のシートから拾う。
    
    Returns:
        (extracted_data, field_source_map)
        field_source_map = {"support_content_1": "モニタリング報告書R7.11", ...}
        各項目がどのシートから取れたかを記録
    """
    if not monitoring_sheets:
        return {}, {}
    
    # すべてのフィールドを空で初期化
    extracted = {key: None for key in EXTRACTION_MAP.keys()}
    field_source = {}  # どのシートから取得したかの記録
    
    # 新しい順に見ていく
    for era_year, month, sheet_name in monitoring_sheets:
        ws = workbook[sheet_name]
        sheet_data = extract_from_worksheet(ws, EXTRACTION_MAP)
        
        for key, val in sheet_data.items():
            # まだ値が入っていない項目にだけセット
            if extracted[key] in (None, "") and val not in (None, ""):
                extracted[key] = val
                field_source[key] = sheet_name
        
        # 全部埋まったら終了
        if all(v not in (None, "") for v in extracted.values()):
            break
    
    return extracted, field_source


def extract_usage_weekday(workbook, case_sheets: list) -> tuple:
    """
    ケース記録シートから利用曜日を抽出。
    複数曜日対応（B51, C51, D51, E51）を「、」で連結。
    最新月から順に見て、最初に見つかった値（どれかのセルに値があるシート）を採用。
    
    表記例:
      - "火・午前"           (週1回、午前だけ)
      - "火・午前、日"       (週2回、火曜午前 + 日曜終日)
      - "月・午前、水・午後、金・午前"  (週3回)
    
    Returns:
        (combined_string, source_sheet_name)
    """
    WEEKDAY_CELLS = ["B51", "C51", "D51", "E51"]
    
    for era_year, month, sheet_name in case_sheets:
        ws = workbook[sheet_name]
        parts = []
        for cell_ref in WEEKDAY_CELLS:
            val = _convert_value(ws[cell_ref].value, "string")
            if val:
                parts.append(val)
        if parts:
            return "、".join(parts), sheet_name
    
    return None, None


# ========================================================
# インポート処理本体
# ========================================================

def analyze_xlsx(file_source) -> dict:
    """
    xlsxファイルを解析して、利用者情報を抽出（フォールバック対応版）
    """
    result = {
        "status": "error",
        "latest_sheet_name": None,
        "latest_period": None,
        "extracted": {},
        "field_sources": {},       # 各項目がどのシート由来か
        "missing_fields": [],
        "warnings": [],
        "usage_weekday": None,
        "usage_weekday_source": None,
    }
    
    try:
        if isinstance(file_source, (str, Path)):
            wb = load_workbook(str(file_source), data_only=True)
        elif isinstance(file_source, bytes):
            wb = load_workbook(BytesIO(file_source), data_only=True)
        else:
            wb = load_workbook(file_source, data_only=True)
    except Exception as e:
        result["warnings"].append(f"ファイル読み込みエラー: {str(e)}")
        return result
    
    # モニタリング報告書シート一覧
    monitoring_sheets = list_monitoring_sheets_newest_first(wb)
    if not monitoring_sheets:
        result["warnings"].append("モニタリング報告書のシートが見つかりません")
        return result
    
    # 最新月の情報（表示用）
    latest_era, latest_month, latest_name = monitoring_sheets[0]
    result["latest_sheet_name"] = latest_name
    result["latest_period"] = {"era_year": latest_era, "month": latest_month}
    
    # モニタリングシートから抽出（フォールバック対応）
    extracted, field_sources = extract_with_fallback(wb, monitoring_sheets)
    result["extracted"] = extracted
    result["field_sources"] = field_sources
    
    # ケース記録から利用曜日を抽出
    case_sheets = list_case_sheets_newest_first(wb)
    weekday, weekday_source = extract_usage_weekday(wb, case_sheets)
    result["usage_weekday"] = weekday
    result["usage_weekday_source"] = weekday_source
    # extractedにも入れる（UI統一のため）
    result["extracted"]["usage_weekday"] = weekday
    if weekday_source:
        result["field_sources"]["usage_weekday"] = weekday_source
    
    # 空欄の項目
    for k, v in result["extracted"].items():
        if v in (None, ""):
            result["missing_fields"].append(k)
    
    # 基本チェック
    if not extracted.get("user_name"):
        result["warnings"].append("利用者氏名が取得できませんでした")
    
    # フォールバックが発生した項目を警告として出す
    fallback_fields = []
    for key, src in field_sources.items():
        if src != latest_name:
            fallback_fields.append((key, src))
    if fallback_fields:
        result["warnings"].append(
            f"{len(fallback_fields)}件の項目は過去月のシートから取得しました"
        )
    
    result["status"] = "success"
    return result


def convert_to_care_plan(extracted: dict, chart_number: str = None) -> dict:
    """抽出データを patient_care_plans テーブル用に変換"""
    care_plan = {
        "chart_number": chart_number or extracted.get("user_name", ""),
        "user_name": extracted.get("user_name", ""),
        # 長期目標
        "long_goal_function":      extracted.get("long_goal_function") or "",
        "long_goal_activity":      extracted.get("long_goal_activity") or "",
        "long_goal_participation": extracted.get("long_goal_participation") or "",
        "long_goal_period_from":   extracted.get("long_goal_period_from"),
        "long_goal_period_to":     extracted.get("long_goal_period_to"),
        # 短期目標
        "short_goal_function":      extracted.get("short_goal_function") or "",
        "short_goal_activity":      extracted.get("short_goal_activity") or "",
        "short_goal_participation": extracted.get("short_goal_participation") or "",
        "short_goal_period_from":   extracted.get("short_goal_period_from"),
        "short_goal_period_to":     extracted.get("short_goal_period_to"),
        # 支援内容
        "support_content_1": extracted.get("support_content_1") or "",
        "support_content_2": extracted.get("support_content_2") or "",
        "support_content_3": extracted.get("support_content_3") or "",
        "support_content_4": extracted.get("support_content_4") or "",
        # 利用曜日（新規）
        "usage_weekday": extracted.get("usage_weekday") or "",
    }
    return care_plan