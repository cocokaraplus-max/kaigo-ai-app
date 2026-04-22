"""
実様式対応 テンプレ差し込み処理
  - 固定値の自動挿入
  - AI生成値の差し込み
  - 既存の罫線・結合・フォントを完全保持
  - 壊れた #REF! 数式は上書きされる（逆にそれで綺麗になる）
"""
import json
import os
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook


BASE = Path(__file__).parent
DATA_DIR = BASE / "data"
OUTPUT_DIR = BASE / "output_real"
DATA_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

RECORDS_JSON = DATA_DIR / "real_records.jsonl"


def fill_template(mapping: dict, data: dict, template_path: str, output_path: str) -> str:
    """テンプレートに全値を差し込む"""
    wb = load_workbook(template_path)
    ws = wb[mapping["sheet_name"]]

    # 1. 固定値（事業所情報等）
    for cell, value in mapping.get("fixed_values", {}).items():
        if cell.startswith("_"):
            continue
        ws[cell] = value

    # 2. 自動値（作成日、年齢）
    auto = mapping.get("auto_values", {})
    if "create_date" in auto:
        ws[auto["create_date"]["cell"]] = date.today()
    if "age" in auto and data.get("_auto_age"):
        ws[auto["age"]["cell"]] = data["_auto_age"]

    # 3. フィールド値
    for key, info in mapping["fields"].items():
        value = data.get(key, "")
        if value in ("", None):
            continue

        cell = info["cell"]
        # 型変換
        if info["type"] == "number":
            try:
                value = float(value) if "." in str(value) else int(value)
            except (ValueError, TypeError):
                pass
        elif info["type"] == "date":
            try:
                for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
                    try:
                        value = datetime.strptime(str(value).strip(), fmt).date()
                        break
                    except ValueError:
                        continue
            except Exception:
                pass

        ws[cell] = value

    wb.save(output_path)
    return output_path


def save_structured_record(mapping_id: str, data: dict) -> dict:
    """JSON Lines 形式で蓄積"""
    record = {
        "mapping_id": mapping_id,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "data": {k: (v.isoformat() if isinstance(v, (date, datetime)) else v) for k, v in data.items()},
    }
    with open(RECORDS_JSON, "a", encoding="utf-8") as f:
        f.write(json.dumps(record, ensure_ascii=False) + "\n")
    return record


def process_case_record(raw_input: str, mapping_path: str = "real_mapping.json") -> dict:
    from ai_generator_real import generate_case_record

    with open(mapping_path, encoding="utf-8") as f:
        mapping = json.load(f)

    structured = generate_case_record(raw_input, mapping)

    base_dir = Path(mapping_path).parent
    template_path = base_dir / mapping["template_file"]
    client_name = structured.get("client_name", "unknown").replace(" ", "") or "unknown"
    date_str = date.today().strftime("%Y-%m-%d")
    output_path = OUTPUT_DIR / f"monitoring_{client_name}_{date_str}.xlsx"

    fill_template(mapping, structured, str(template_path), str(output_path))
    record = save_structured_record(mapping["template_id"], structured)

    return {
        "structured_data": structured,
        "output_file": str(output_path),
        "record": record,
    }


if __name__ == "__main__":
    sample = """
2026-04-19 山田花子さん訪問
担当者: 鈴木一郎
フリガナ: ヤマダハナコ
生年月日: 1945-03-15
性別: 女
要介護度: 要介護2
体重: 48.5
歩行距離が延びてきており、笑顔も増えた。機能訓練にも意欲的。
新しい希望: 外出支援も受けたい
"""
    result = process_case_record(sample)
    print(f"出力ファイル: {result['output_file']}")
    print(f"\n--- 構造化データ（抜粋） ---")
    for k, v in result['structured_data'].items():
        if v and not k.startswith("_"):
            print(f"  {k}: {v}")
